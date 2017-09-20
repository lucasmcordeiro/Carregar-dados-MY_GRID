import openpyxl as xl
import sys
from collections import OrderedDict
from mygrid.rede import Subestacao, Alimentador, Setor, Chave
from mygrid.rede import Trecho, NoDeCarga, Transformador, Condutor
from mygrid.util import Fasor
from bs4 import BeautifulSoup as bs

def carregar_dados(arquivo):
    wb = xl.load_workbook(arquivo)

    trechos_sheet = wb.get_sheet_by_name('trechos')
    chaves_sheet=wb.get_sheet_by_name('chaves')
    a=1
    dados = OrderedDict()
    trechos = list()
    for i in range(2, trechos_sheet.max_row+1):
        if trechos_sheet.cell(row=i, column=1).value is not None:
            a+=1
    for i in range(2, trechos_sheet.max_row+1):
        trecho = OrderedDict()
        if a==trechos_sheet.max_row:
            trecho['nome'] = str(trechos_sheet.cell(row=i, column=1).value)
        elif a!=1:
            print "A coluna trechos nao foi preenchida totalmente. Preencha totalmente a coluna trechos ou a deixe totalmente vazia."
            sys.exit()
            
        else:
            trecho['nome'] = str(trechos_sheet.cell(row=i, column=2).value)+'_'+str(trechos_sheet.cell(row=i, column=3).value)
        trecho['n1'] = str(trechos_sheet.cell(row=i, column=2).value)
        trecho['n2'] = str(trechos_sheet.cell(row=i, column=3).value)
        trecho['setor'] = str(trechos_sheet.cell(row=i, column=5).value)
        trecho['comprimento'] = float(trechos_sheet.cell(row=i, column=6).value)
        trecho['alimentador'] = str(trechos_sheet.cell(row=i, column=4).value)
        trecho['condutores'] = str(trechos_sheet.cell(row=i, column=7).value)
        trechos.append(trecho)

    nos_sheet = wb.get_sheet_by_name('nos')

    nos = list()
    for i in range(2, nos_sheet.max_row+1):
        no = OrderedDict()
        no['nome'] = str(nos_sheet.cell(row=i, column=1).value)
        no['P'] = float(nos_sheet.cell(row=i, column=2).value)
        no['Q'] = float(nos_sheet.cell(row=i, column=3).value)
        nos.append(no)

    chaves=list()
    for i in range(2,chaves_sheet.max_row+1):
        chaves.append(str(chaves_sheet.cell(row=i,column=1).value))

    dados['trechos'] = trechos
    dados['nos'] = nos
    dados['chaves'] = chaves
    

    return dados


def gerar_chaves(arquivo): # Gera os objetos chaves
    
    wb = xl.load_workbook(arquivo)
    chaves_sheet=wb.get_sheet_by_name('chaves')

    chaves = list()
    for i in range(2,chaves_sheet.max_row+1):
        chave = OrderedDict()
        nome=str(chaves_sheet.cell(row=i, column=1).value)
        estado=int(chaves_sheet.cell(row=i, column=2).value)
        chave=Chave(nome=nome,estado=estado)
        chaves.append(chave)

    return chaves


def gerar_nos_de_carga(vizinhos,dados): #Gera os objetos NoDeCarga
    potencia=dados['nos']
    chaves=dados['chaves']
    nos_de_cargas=list()
    chaves_vizinhas=list()

    for i in range(len(vizinhos)):
        chave_vizinha=OrderedDict()
        chave_vizinha['nome']=vizinhos[i]['no']
        chave_vizinha['vizinhos']=list()
        for j in range(len(vizinhos[i]['vizinhos'])):
            if vizinhos[i]['vizinhos'][j] in chaves:
                chave_vizinha['vizinhos'].append(vizinhos[i]['vizinhos'][j])
        chaves_vizinhas.append(chave_vizinha)



    for i in range(len(vizinhos)):
        nos_de_carga=OrderedDict()
        nome=vizinhos[i]['no'].upper()
        real= potencia[i]['P']*1e3
        img=potencia[i]['Q']*1e3
        vizinhos_i=[]
        chaves_i=chaves_vizinhas[i]['vizinhos']
        for j in range(len(vizinhos[i]['vizinhos'])):
            if vizinhos[i]['vizinhos'][j] not in chaves:
                vizinhos_i.append(vizinhos[i]['vizinhos'][j].upper())
            else:
                for x in range(len(vizinhos)):
                    if vizinhos[i]['vizinhos'][j] in vizinhos[x]['vizinhos'] and x!=i:
                         vizinhos_i.append(vizinhos[x]['no'].upper())


        nos_de_carga=NoDeCarga(nome=nome,vizinhos=vizinhos_i,potencia=Fasor(real=real, imag=img, tipo=Fasor.Potencia), chaves=chaves_i)
        nos_de_cargas.append(nos_de_carga)
    return nos_de_cargas

def gerar_setores(dados,nos_de_cargas): #Gera os objetos do tipo Setor
    setor_dados=dados['trechos']
    setor_nome=list()
    chaves_nome=dados['chaves']
    for i in range(len(setor_dados)):
        if setor_dados[i]['setor'] not in setor_nome:
            setor_nome.append(setor_dados[i]['setor'])

    setor_ii=list()
    for i in range(len(setor_nome)):
        setor_i=OrderedDict()
        setor_i['nome']=setor_nome[i]
        setor_i['nos_contidos_setor']=list()
        for j in range(len(setor_dados)):
            if setor_nome[i]==setor_dados[j]['setor']:
                if setor_dados[j]['n1'] not in setor_i['nos_contidos_setor']:
                    setor_i['nos_contidos_setor'].append(setor_dados[j]['n1'])
                if setor_dados[j]['n2'] not in setor_i['nos_contidos_setor']:
                    setor_i['nos_contidos_setor'].append(setor_dados[j]['n2'])
        setor_ii.append(setor_i)

    setores=list()
    for i in range(len(setor_ii)):
        setores_i=OrderedDict()
        nome=setor_ii[i]['nome'].upper();
        vizinhos=[]
        nos_cargas=[]
        nos=[]
        for j in range(len(setor_ii)):
            if setor_ii[i]['nome']!=setor_ii[j]['nome']:
                for y in range(len(setor_ii[i]['nos_contidos_setor'])):
                    if setor_ii[i]['nos_contidos_setor'][y] not in chaves_nome and setor_ii[i]['nos_contidos_setor'][y] not in nos :
                        nos.append(setor_ii[i]['nos_contidos_setor'][y])

                    if setor_ii[i]['nos_contidos_setor'][y] in setor_ii[j]['nos_contidos_setor']:
                        vizinhos.append(setor_ii[j]['nome'].upper())
        
        for j in range(len(nos)):
            for x in range(len(nos_de_cargas)):
                if nos[j].upper()==nos_de_cargas[x].nome:
                    nos_cargas.append(nos_de_cargas[x])
        setores_i=(Setor(nome=nome,vizinhos=vizinhos, nos_de_carga=nos_cargas))
        setores.append(setores_i)
    return setores
    
def gerar_alimentadores(dados,arquivo,trechos_d,setores,chaves): # Gera objetos do tipo Alimentador
    wb = xl.load_workbook(arquivo)
    alimentadores_sheet = wb.get_sheet_by_name('alimentadores')
    trechos_alimentadores=wb.get_sheet_by_name('trechos')
    trechos=dados['trechos']
    chaves_nome=dados['chaves']
    ali=list()
    for i in range(2,alimentadores_sheet.max_row+1):
        ali_i=OrderedDict()
        ali_i['alimentador']=(str(alimentadores_sheet.cell(row=i,column=1).value))
        ali_i['sub_est']=(str(alimentadores_sheet.cell(row=i,column=2).value))
        ali.append(ali_i)
        
    alimentador=list()
    for i in range(len(ali)):
        alimentadores_i=OrderedDict()
        alimentadores_i['nome']=ali[i]['sub_est']+'_'+ali[i]['alimentador']
        alimentadores_i['setores']=list()
        alimentadores_i['trechos']=list()
        alimentadores_i['chaves']=list()
        for j in range(2,trechos_alimentadores.max_row+1):
            if ali[i]['alimentador']==str(trechos_alimentadores.cell(row=j,column=4).value):
                if str(trechos_alimentadores.cell(row=j,column=5).value) not in alimentadores_i['setores']:
                    alimentadores_i['setores'].append(str(trechos_alimentadores.cell(row=j,column=5).value))
                if trechos[j-2]['nome'] not in  alimentadores_i['trechos']:
                    alimentadores_i['trechos'].append(trechos[j-2]['nome'])
                if str(trechos_alimentadores.cell(row=j,column=2).value) in chaves_nome:
                    if str(trechos_alimentadores.cell(row=j,column=2).value) not in alimentadores_i['chaves']:
                        alimentadores_i['chaves'].append(str(trechos_alimentadores.cell(row=j,column=2).value))
                if str(trechos_alimentadores.cell(row=j,column=3).value) in chaves_nome:
                    if str(trechos_alimentadores.cell(row=j,column=3).value) not in alimentadores_i['chaves']:
                        alimentadores_i['chaves'].append(str(trechos_alimentadores.cell(row=j,column=3).value))
               
        alimentador.append(alimentadores_i)
    
    alimentadores=list()
    for i in range(len(alimentador)):
        nome=alimentador[i]['nome'].upper()
        setor=list()
        trecho1=list()
        chave=list()
        alimentadores_ii=OrderedDict()
        for j in range(len(alimentador[i]['setores'])):
            for x in range(len(setores)):
                if alimentador[i]['setores'][j].upper()==setores[x].nome:
                    setor.append(setores[x])

        for j in range(len(alimentador[i]['trechos'])):
            for x in range(len(trechos_d)):
                if alimentador[i]['trechos'][j].upper()==trechos_d[x].nome:
                    trecho1.append(trechos_d[x])

        for j in range(len(alimentador[i]['chaves'])):
            for x in range(len(chaves)):
                if alimentador[i]['chaves'][j]==chaves[x].nome:
                    chave.append(chaves[x])
        alimentadores_ii=Alimentador(nome=nome,setores=setor,trechos=trecho1,chaves=chave)
        alimentadores.append(alimentadores_ii)
    return alimentadores

def gerar_trechos(dados,nos_de_carga,chaves): #Gera os Objetos do tipo Trecho
    trechos_lista=dados['trechos']
    trechos=list()

    cabos_xml=bs(open('cabos.xml','r'),'xml')
    cabos=cabos_xml.find_all('condutor')
    for i in range(len(trechos_lista)):
        trecho=OrderedDict()
        
        for j in range(len(nos_de_carga)):

            if trechos_lista[i]['n1'].upper()==nos_de_carga[j].nome:
                n1=nos_de_carga[j]
            if trechos_lista[i]['n2'].upper()==nos_de_carga[j].nome:
                n2=nos_de_carga[j]

        for j in range(len(chaves)):
            if trechos_lista[i]['n1']==chaves[j].nome:
                n1=chaves[j]
            if trechos_lista[i]['n2']==chaves[j].nome:
                n2=chaves[j]
    
        for cab in cabos:
            if trechos_lista[i]['condutores']==cab.get('nome'):
                cond_1=Condutor(nome=cab.get('nome'),
                                rp=float(cab.get('rp')),
                                rz=float(cab.get('rz')),
                                xp=float(cab.get('xp')),
                                xz=float(cab.get('xz')),
                                ampacidade=float(cab.get('ampacidade')))
        nome=trechos_lista[i]['nome'].upper()
        comprimento=trechos_lista[i]['comprimento']
        trecho=Trecho(nome=nome,n1=n1,n2=n2,condutor=cond_1,comprimento=comprimento)
        trechos.append(trecho)
    return trechos
def gerar_ligacao_chaves_setores(dados,chaves,setores): # Faz a ligacao das chaves com seus respectivos setores
    trechos=dados['trechos']
    chaves_nome=dados['chaves']
    ligacao_chave=list()
    ch_exist=[]
    for i in range(len(trechos)):
        ligacao=OrderedDict()
        n_1='nope'
        n_2='nope'

        if trechos[i]['n1'] in chaves_nome and trechos[i]['n1'] not in ch_exist :
            ch_exist.append(trechos[i]['n1'])
            n_1=trechos[i]['n1']
            ligacao['nome_chave']=n_1
            ligacao['n1']=trechos[i]['setor']

            for j in range(len(trechos)-(1+i)):
                if trechos[j+1+i]['n1']==n_1:
                    ligacao['n2']=trechos[j+1+i]['setor']
                elif trechos[j+1+i]['n2'][0:2]==n_1:
                    ligacao['n2']=trechos[j+1+i]['setor']

        elif trechos[i]['n2'] in chaves_nome and trechos[i]['n2'] not in ch_exist :
            ch_exist.append(trechos[i]['n2'])
            n_2=trechos[i]['n2']
            ligacao['nome_chave']=n_2
            ligacao['n1']=trechos[i]['setor']

            for j in range(len(trechos)-(1+i)):
                if trechos[j+1+i]['n1']==n_2:
                    ligacao['n2']=trechos[j+1+i]['setor']
                elif trechos[j+1+i]['n2']==n_2:
                    ligacao['n2']=trechos[j+1+i]['setor']

        if n_1!='nope' or n_2!='nope':
            ligacao_chave.append(ligacao)
    
    for i in range(len(ligacao_chave)):
        for j in range(len(chaves)):
            if ligacao_chave[i]['nome_chave']==chaves[j].nome:
                for x in range(len(setores)):
                    if  ligacao_chave[i]['n1'].upper()==setores[x].nome:
                        chaves[j].n1=setores[x]
                for x in range(len(setores)):
                    if  ligacao_chave[i]['n2'].upper()==setores[x].nome:
                        chaves[j].n2=setores[x]
                    
    return chaves

def _identificar_nos_vizinhos(dados): # identifica os nos vizinhos
    trechos = dados['trechos']
    nos = dados['nos']

    vizinhos = list()
    for no in nos:
        vizinhanca = OrderedDict()
        vizinhanca['no'] = no['nome']
        vizinhanca['vizinhos'] = list()
        for trecho in trechos:
            if trecho['n1'] == no['nome']:
                vizinhanca['vizinhos'].append(trecho['n2'])
            elif trecho['n2'] == no['nome']:
                vizinhanca['vizinhos'].append(trecho['n1'])
        vizinhos.append(vizinhanca)

    return vizinhos

def gerar_transformadores(arquivo): # gera os objetos do tipo Transformador
    wb=xl.load_workbook(arquivo)
    transformadores_sheet=wb.get_sheet_by_name('transformadores')
    transformadores=list()
    for i in range(2,transformadores_sheet.max_column+1):
        transformador=OrderedDict()
        nome=str(transformadores_sheet.cell(row=10,column=i).value).upper()+'_'+str(transformadores_sheet.cell(row=1,column=i).value).upper()
        #float(trechos_sheet.cell(row=i, column=6).value)
        tensao_primario_mod=float(transformadores_sheet.cell(row=2,column=i).value)*1e3
        tensao_primario_ang=float(transformadores_sheet.cell(row=3,column=i).value)
        tensao_secundario_mod=float(transformadores_sheet.cell(row=4,column=i).value)*1e3
        tensao_secundario_ang=float(transformadores_sheet.cell(row=5,column=i).value)
        potencia_mod=float(transformadores_sheet.cell(row=6,column=i).value)*1e6
        potencia_ang=float(transformadores_sheet.cell(row=7,column=i).value)
        impedancia_real=float(transformadores_sheet.cell(row=8,column=i).value)
        impedancia_imag=float(transformadores_sheet.cell(row=9,column=i).value)
        transformador=Transformador(nome=nome,
                                    tensao_primario=Fasor(mod=tensao_primario_mod,ang=tensao_primario_ang,tipo=Fasor.Tensao),
                                    tensao_secundario=Fasor(mod=tensao_secundario_mod,ang=tensao_secundario_ang,tipo=Fasor.Tensao),
                                    potencia=Fasor(mod=potencia_mod,ang=potencia_ang,tipo=Fasor.Potencia),
                                    impedancia=Fasor(real=impedancia_real,imag=impedancia_imag,tipo=Fasor.Impedancia))
        transformadores.append(transformador)
    return transformadores

def gerar_sub_estacao(alimentadores,arquivo,transformadores): #gera os objetos do tipo Subestacao
    wb = xl.load_workbook(arquivo)
    alimentadores_sheet = wb.get_sheet_by_name('alimentadores')
    
    lista_nome_sub_est=list()
    for i in range(len(alimentadores)):
        nome_sub_est=''
        for j in range(len(alimentadores[i].nome)):
            if alimentadores[i].nome[j]!='_':
                nome_sub_est=nome_sub_est+(alimentadores[i].nome[j])
            else:
                break
        lista_nome_sub_est.append(nome_sub_est)

    subestacao=list()
    for i in range(len(lista_nome_sub_est)):
        sub=OrderedDict()
        aliment=list()
        for j in range(len(alimentadores)):
            a=0
            for x in alimentadores[j].nome:
                if x != '_':
                    a+=1
                else:
                    break
            if lista_nome_sub_est[i]==alimentadores[j].nome[0:a]:
                nome=lista_nome_sub_est[i]
                aliment.append(alimentadores[j])
                print nome
        transformador=[]
        for j in range(len(transformadores)):
            a=0
            for x in transformadores[j].nome:
                if x != '_':
                    a+=1
                else:
                    break
            if nome==transformadores[j].nome[0:a]:
                print nome,transformadores[j].nome[0:a]
                transformador.append(transformadores[j])
        sub=Subestacao(nome=nome,alimentadores=aliment,transformadores=transformador)
        subestacao.append(sub)

    lista_nome_alimentadores=list()
    for i in range(len(alimentadores)):
        p={alimentadores[i].nome:alimentadores[i]}
        lista_nome_alimentadores.append(p)
    
    for i in range(len(alimentadores)):
        for j in range(2,alimentadores_sheet.max_row+1):
            if subestacao[i].nome == str(alimentadores_sheet.cell(row=j,column=2).value).upper():
                alimentadores[i].ordenar(raiz=str(alimentadores_sheet.cell(row=j,column=3).value).upper())

    arvore_subestacao=list()
    for i in range(len(alimentadores)):
        p=alimentadores[i].gerar_arvore_nos_de_carga()
        arvore_subestacao.append(p)


    return subestacao

if __name__ == '__main__':
    arquivo=str(raw_input('Entre com o nome do arquivo xlsx a ser carregado: '))+'.xlsx'
    dados = carregar_dados(arquivo)
    chaves=gerar_chaves(arquivo)
    vizinhos = _identificar_nos_vizinhos(dados)
    nos_de_carga=gerar_nos_de_carga(vizinhos,dados)
    nos=dados['trechos']
    nos_de_carga=gerar_nos_de_carga(vizinhos,dados)
    trechos=gerar_trechos(dados,nos_de_carga,chaves)
    setores=gerar_setores(dados,nos_de_carga)
    gerar_ligacao_chaves_setores(dados,chaves,setores)
    alimentadores=gerar_alimentadores(dados,arquivo,trechos,setores,chaves)
    transformadores=gerar_transformadores(arquivo)
    subestacoes=gerar_sub_estacao(alimentadores,arquivo,transformadores)

    no={}
    for i in range(len(nos_de_carga)):
        no[nos_de_carga[i].nome]=nos_de_carga[i]
    chave={}
    for i in range(len(chaves)):
        chave[chaves[i].nome]=chaves[i]
    trecho={}
    for i in range(len(trechos)):
        trecho[trechos[i].nome]=trechos[i]
    setor={}
    for i in range(len(setores)):
        setor[setores[i].nome]=setores[i]
    alimentador={}
    for i in range(len(alimentadores)):
        alimentador[alimentadores[i].nome]=alimentadores[i]
    subestacao={}
    for i in range(len(subestacoes)):
        subestacao[subestacoes[i].nome]=subestacoes[i]

    
   